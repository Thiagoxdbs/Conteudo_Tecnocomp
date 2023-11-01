import aspose.words as aw
from docx import Document
#import tkinter as tk #USADO PARA CRIAR A INTERFACE DO APP
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


banco = []

#Alex Moreira De Lima
#Backup ok
dados0 = {'Nome_Completo':'Alex Moreira De Lima'
,'CPF':'22522027859'
,'RG':'29293581-x'
,'Profissão':'Supervisor de Logística'
,'Setor':'Logistica ADM'
,'Quantidade':'1'
,'Marca_antigo':'Samsung'
,'Modelo_antigo':'Galaxy A12'
,'Marca_novo':'Samsung'
,'Modelo_novo':'Galaxy A53'
,'carregador':'Nâo'
,'IMEI_antigo':'359410827048784'
,'IMEI_novo':'350603383809362'
,'Telefone':'\(11\)978639549'
,'Obs_novo':'Aparelho entregue com carregador, cabo USB, extrator de chip, pelicula, capa de silicone e manual do usuário.'
,'Obs_antigo':'Sem carregador, está sem avaria na tela.'}

dados1 = {'Nome_Completo':'Andreza de Andrade Oliveira de Santi'
,'CPF':'20510371833'
,'RG':'273241552'
,'Profissão':'Coordenadora de Ser. Odontologico'
,'Setor':'Administração Odontológica'
,'Quantidade':'1'
,'Marca_antigo':'LG'
,'Modelo_antigo':'K10'
,'Marca_novo':'Samsung'
,'Modelo_novo':'Galaxy A53'
,'carregador':'Não'
,'IMEI_antigo':'35566909328795405'
,'IMEI_novo':'350603383815609'
,'Telefone':'(11)975985606'
,'Obs_novo':'Aparelho entregue com carregador, cabo USB, extrator de chip, pelicula, capa de silicone e manual do usuário.'
,'Obs_antigo':'Não recebeu carregador, está sem avaria na tela.'}

#Isabel Cristina Ruas de Azevedo de Oliveira
dados2 = {'Nome_Completo':'Isabel Cristina Ruas de Azevedo de Oliveira'
,'CPF':'29234463803'
,'RG':'328065195'
,'Profissão':''
,'Setor':''
,'Quantidade':''
,'Marca_antigo':''
,'Modelo_antigo':''
,'Marca_novo':''
,'Modelo_novo':''
,'carregador':''
,'IMEI_antigo':''
,'IMEI_novo':'350603383809180'
,'Telefone':'11957721389'
,'Obs_novo':''
,'Obs_antigo':''}

#Denis Martins de Carvalho
dados3 = {'Nome_Completo':'Denis Martins de Carvalho'
,'CPF':''
,'RG':''
,'Profissão':''
,'Setor':''
,'Quantidade':''
,'Marca_antigo':''
,'Modelo_antigo':''
,'Marca_novo':''
,'Modelo_novo':''
,'carregador':''
,'IMEI_antigo':''
,'IMEI_novo':''
,'Telefone':''
,'Obs_novo':''
,'Obs_antigo':''}


#Ana Neri Mulinari
#Backup ok
dados4 = {'Nome_Completo':'Ana Neri Mulinari'
,'CPF':'16353330839'
,'RG':'272465598'
,'Profissão':'Gerente Técnica Operacional'
,'Setor':'Regulamentação Médica'
,'Quantidade':'1'
,'Marca_antigo':'Samsung'
,'Modelo_antigo':'Galaxy A12'
,'Marca_novo':'Samsung'
,'Modelo_novo':'Galaxy A53'
,'carregador':'Sim'
,'IMEI_antigo':'359410827013200'
,'IMEI_novo':'350603383816003'
,'Telefone':'(11)978639525'
,'Obs_novo':'Aparelho entregue com carregador, cabo USB, extrator de chip, pelicula, capa de silicone e manual do usuário.'
,'Obs_antigo':'Devolveu com carregador, com caixa do aparelho, com capa e está sem avaria na tela.'}


banco.append(dados0)
banco.append(dados1)
banco.append(dados2)
banco.append(dados3)
banco.append(dados4)

wb = load_workbook('')
Print(wb)

print(banco)
#Abrindo arquivo que vai ser editado
def termo(arquivo):
	documento = Document(arquivo)
	#Trocando caracteres no texto do word
	for c in range(0,4):
		for paragrafo in documento.paragraphs:	
			paragrafo.text = paragrafo.text.replace("nome_t",f"{banco[c]['Nome_Completo']}")
			paragrafo.text = paragrafo.text.replace("rg_t" , banco[c]['RG'])
			paragrafo.text = paragrafo.text.replace("cpf_t" , banco[c]['CPF'])
			paragrafo.text = paragrafo.text.replace("profissao_t" , banco[c]['Profissão'])
			paragrafo.text = paragrafo.text.replace("unidade_t" , banco[c]['Quantidade'])
			paragrafo.text = paragrafo.text.replace("marca_tn" , banco[c]['Quantidade'])
			paragrafo.text = paragrafo.text.replace("modelo_tn" , banco[c]['Modelo_novo'])
			paragrafo.text = paragrafo.text.replace("telefone_t" , banco[c]['Telefone'])
			paragrafo.text = paragrafo.text.replace("emei_tn" , banco[c]['IMEI_novo'])
			paragrafo.text = paragrafo.text.replace("setor_t" , banco[c]['Setor'])
			paragrafo.text = paragrafo.text.replace("obs_tn" , banco[c]['Obs_novo'])
			paragrafo.text = paragrafo.text.replace("Marca_ta" , banco[c]['Marca_antigo'])
			paragrafo.text = paragrafo.text.replace("Modelo_ta" , banco[c]['Modelo_antigo'])
			paragrafo.text = paragrafo.text.replace("imei_ta" , banco[c]['IMEI_antigo'])
			paragrafo.text = paragrafo.text.replace("obs_ta" , banco[c]['Obs_antigo'])
			paragrafo.text = paragrafo.text.replace("carregador_t" , banco[c]['carregador'])
			
			#Salvando conteudo em um novo arquivo word	
			documento.save(f"{ banco[c]['Nome_Completo']}- {arquivo}")
			print('TERMO FOI GERADO')

termo('Entrega_Smartphone.docx')
termo('Recebimento_Smartphone.docx')
