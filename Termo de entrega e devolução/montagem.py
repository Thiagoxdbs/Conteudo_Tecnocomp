from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook(filename = 'montagem.xlsx')
sheet_ranges = wb['nome']
print(sheet_ranges['A2'].value)